from fastapi import FastAPI, Depends, HTTPException, status, Query, Request, UploadFile, File, Body
# NOTA: CORSMiddleware Starlette 1.3.1 digantikan dengan custom middleware di bawah
#       (bug: Access-Control-Allow-Origin tidak dihantar bila allow_credentials=True)
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
from backend.database import get_db, init_db
from backend.auth import get_current_user
from backend.secure_auth import (
    hash_kata_laluan, sahkan_kata_laluan, create_access_token,
    get_pengguna_dari_db, login_endpoint
)
from backend.config import settings
import sqlite3
import pandas as pd
import openpyxl
from io import BytesIO
import re
import json
import random
from datetime import datetime
import os

app = FastAPI(title="Sistem Pengurusan Pengundi Parlimen P170 Tuaran")

# ===== CUSTOM CORS MIDDLEWARE =====
# NOTA: Starlette 1.3.1 CORSMiddleware mempunyai bug:
#       Access-Control-Allow-Origin tidak dihantar bila allow_credentials=True.
#       Guna @app.middleware("http") sebagai ganti.
#       Rujukan: https://github.com/encode/starlette/issues/1172

@app.middleware("http")
async def custom_cors_middleware(request: Request, call_next):
    """CORS middleware yang memastikan Access-Control-Allow-Origin sentiasa ada."""

    origin = request.headers.get("origin", "")
    allowed = settings.cors_origins

    # Tentukan ACAO: echo balik origin jika dibenarkan
    if origin and origin in allowed:
        acao = origin
    elif origin:
        acao = origin  # echo back walaupun tak dalam senarai (untuk debugging)
    else:
        acao = allowed[0] if allowed else "*"

    # OPTIONS preflight — return 200 terus dengan CORS headers
    if request.method == "OPTIONS":
        return JSONResponse(
            content={},
            status_code=200,
            headers={
                "Access-Control-Allow-Origin": acao,
                "Access-Control-Allow-Credentials": "true",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, PATCH, OPTIONS",
                "Access-Control-Allow-Headers": "*",
                "Access-Control-Max-Age": "86400",
            }
        )

    # Bukan OPTIONS — proses normal, kemudian tambah CORS headers
    response = await call_next(request)
    response.headers["Access-Control-Allow-Origin"] = acao
    response.headers["Access-Control-Allow-Credentials"] = "true"
    return response


# Static files path (digunakan oleh catch-all route di bawah)
DEV_STATIC_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'frontend')
PROD_STATIC_PATH = settings.STATIC_DIR

if settings.PRODUCTION:
    STATIC_ROOT = PROD_STATIC_PATH
    print(f"✅ PRODUCTION mode — akan serve static dari: {STATIC_ROOT}")
else:
    STATIC_ROOT = DEV_STATIC_PATH
    print(f"✅ DEVELOPMENT mode — akan serve static dari: {STATIC_ROOT}")

if not os.path.exists(STATIC_ROOT):
    print(f"⚠️  WARNING: Static directory tidak wujud: {STATIC_ROOT}")


# ===== GLOBAL EXCEPTION HANDLER =====
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    import sys
    print(f"❌ GLOBAL EXCEPTION HANDLER: {type(exc).__name__}: {str(exc)}", file=sys.stderr)
    traceback.print_exc(file=sys.stderr)
    return JSONResponse(
        status_code=500,
        content={"error": "Internal Server Error", "details": str(exc), "type": type(exc).__name__}
    )


# ===== MODEL DATA =====
class LoginRequest(BaseModel):
    username: str
    kata_laluan: str


class PengundiCreate(BaseModel):
    no_kp: str
    nama_penuh: str
    jantina: Optional[str] = None
    tahun_lahir: Optional[int] = None
    dun: Optional[str] = None  # Kod DUN (contoh: "N12") — akan resolve ke dun_id
    dm: str
    lokaliti: Optional[str] = None
    no_telefon: Optional[str] = None
    status_sokongan: Optional[str] = None
    status_fizikal: Optional[str] = "Hidup"
    ketua_keluarga_id: Optional[int] = None
    pegawai_penyelaras_id: Optional[int] = None
    ketua_keluarga_nama_baru: Optional[str] = None
    pegawai_penyelaras_nama_baru: Optional[str] = None


class PengundiUpdate(BaseModel):
    nama_penuh: Optional[str] = None
    jantina: Optional[str] = None
    tahun_lahir: Optional[int] = None
    dm: Optional[str] = None
    lokaliti: Optional[str] = None
    no_telefon: Optional[str] = None
    status_sokongan: Optional[str] = None
    status_fizikal: Optional[str] = None
    adalah_pemilik_apps: Optional[bool] = None
    ketua_keluarga_id: Optional[int] = None
    pegawai_penyelaras_id: Optional[int] = None


class PenggunaCreate(BaseModel):
    username: str
    nama_penuh: str
    kata_laluan: str
    peranan: str = "Pemerhati"
    dm: Optional[str] = None


# ===== AUDIT TRAIL HELPER =====
def log_activity(request: Request, user: dict, tindakan: str, penerangan: str, no_kp_terlibat: str = None):
    """Log aktiviti pengguna ke dalam audit_logs untuk pematuhan PDPA."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO audit_logs (user_id, username, peranan, tindakan, penerangan, no_kp_terlibat, endpoint, ip_address, user_agent, dicipta_pada)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user.get("user_id"),
            user["username"],
            user["peranan"],
            tindakan,
            penerangan,
            no_kp_terlibat,
            request.url.path if hasattr(request, 'url') else None,
            request.client.host if request.client else "localhost",
            request.headers.get("user-agent") if hasattr(request, 'headers') else None,
            datetime.now().isoformat()
        ))
        db.commit()
        db.close()
    except Exception as e:
        # Kegagalan logging tidak patut menjejaskan request utama (terutamanya di Vercel serverless)
        print(f"⚠️ log_activity gagal (non-critical): {type(e).__name__}: {str(e)}")


# ===== EVENT STARTUP =====
@app.on_event("startup")
def startup():
    try:
        # Redirect stdout to stderr supaya print() tidak cemarkan JSON response
        import sys
        sys.stdout = sys.stderr
        
        init_db()
        
        # Seed default users individually if they don't exist
        db = get_db()
        cursor = db.cursor()
        default_users = [
            ("admin", "Admin Sistem", hash_kata_laluan("admin123"), "Admin"),
            ("petugas", "Petugas Padang", hash_kata_laluan("petugas123"), "Petugas Padang"),
            ("pemerhati", "Pemerhati", hash_kata_laluan("pemerhati123"), "Pemerhati"),
        ]
        now = datetime.now().isoformat()
        seeded = 0
        for username, nama_penuh, kata_laluan, peranan in default_users:
            cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO users (username, nama_penuh, kata_laluan, peranan, dm, aktif, dicipta_pada) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (username, nama_penuh, kata_laluan, peranan, None, 1, now)
                )
                seeded += 1
        if seeded > 0:
            db.commit()
            print(f"✅ {seeded} pengguna lalai telah dicipta")
        else:
            print("✅ Semua pengguna lalai sudah wujud")
        
        # Seed sample pengundi data if database empty
        cursor.execute("SELECT COUNT(*) FROM pengundi")
        pengundi_count = cursor.fetchone()[0]
        if pengundi_count == 0:
            print("📦 Database pengundi kosong - memasukkan data sample...")
            try:
                from backend.seed_data import seed_database
                seed_database()
                print("✅ Data sample pengundi berjaya dimasukkan!")
            except Exception as e:
                print(f"⚠️ Gagal seed data: {e}")
        else:
            print(f"✅ Database pengundi sudah mempunyai {pengundi_count} rekod")
        
        db.close()
    except Exception as e:
        import traceback
        print(f"❌ KRITIKAL: Startup function gagal: {e}", file=__import__('sys').stderr)
        traceback.print_exc()
        print("⚠️ App masih berjalan — endpoint akan return error 500 jika guna database.")


# ===== AUTO-CREATE DUN / PDM =====
class DunCreate(BaseModel):
    kod: str
    nama: str

class PdmCreate(BaseModel):
    nama: str
    dun_kod: str

class LokalitiCreate(BaseModel):
    nama: str
    dm: str  # Nama PDM untuk dikaitkan

@app.post("/api/dun")
def create_dun(request: Request, data: DunCreate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin", "Petugas Padang"])
    db = get_db()
    cursor = db.cursor()
    kod = data.kod.strip().upper()
    nama = data.nama.strip().upper()
    # Check if DUN already exists
    cursor.execute("SELECT id FROM dun WHERE kod = ?", (kod,))
    existing = cursor.fetchone()
    if existing:
        db.close()
        return {"success": True, "id": existing[0], "kod": kod, "nama": nama, "existing": True}
    # Get parlimen_id for P170
    cursor.execute("SELECT id FROM parlimen WHERE kod = 'P170'")
    parlimen = cursor.fetchone()
    if not parlimen:
        db.close()
        raise HTTPException(status_code=500, detail="Parlimen P170 not found")
    parlimen_id = parlimen[0]
    cursor.execute("INSERT INTO dun (parlimen_id, kod, nama) VALUES (?, ?, ?)", (parlimen_id, kod, nama))
    db.commit()
    new_id = cursor.lastrowid
    db.close()
    return {"success": True, "id": new_id, "kod": kod, "nama": nama}

@app.delete("/api/dun/{dun_kod}")
def delete_dun(dun_kod: str, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    
    kod = dun_kod.strip().upper()
    
    # Protect core DUNs (N12-N15)
    if kod in ["N12", "N13", "N14", "N15"]:
        db.close()
        raise HTTPException(status_code=400, detail=f"DUN {kod} adalah teras sistem dan tidak boleh dipadamkan")
    
    # Check if DUN exists
    cursor.execute("SELECT id FROM dun WHERE kod = ?", (kod,))
    dun = cursor.fetchone()
    if not dun:
        db.close()
        raise HTTPException(status_code=404, detail=f"DUN {kod} tidak ditemui")
    
    dun_id = dun[0]
    
    # Check if any pengundi linked to this DUN
    cursor.execute("SELECT COUNT(*) FROM pengundi WHERE dun_id = ?", (dun_id,))
    pengundi_count = cursor.fetchone()[0]
    if pengundi_count > 0:
        db.close()
        raise HTTPException(status_code=400, detail=f"DUN {kod} masih mempunyai {pengundi_count} pengundi. Tidak boleh dipadam.")
    
    # Check if any pdm linked to this DUN
    cursor.execute("SELECT COUNT(*) FROM pdm WHERE dun_id = ?", (dun_id,))
    pdm_count = cursor.fetchone()[0]
    if pdm_count > 0:
        db.close()
        raise HTTPException(status_code=400, detail=f"DUN {kod} masih mempunyai {pdm_count} PDM. Tidak boleh dipadam.")
    
    # Delete the DUN
    cursor.execute("DELETE FROM dun WHERE id = ?", (dun_id,))
    db.commit()
    db.close()
    
    return {"success": True, "message": f"DUN {kod} berjaya dipadamkan"}

@app.post("/api/pdm")
def create_pdm(request: Request, data: PdmCreate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin", "Petugas Padang"])
    db = get_db()
    cursor = db.cursor()
    nama = data.nama.strip().upper()
    dun_kod = data.dun_kod.strip().upper()
    # Check if PDM already exists in same DUN
    cursor.execute("""
        SELECT p.id FROM pdm p 
        JOIN dun d ON d.id = p.dun_id 
        WHERE p.nama = ? AND d.kod = ?
    """, (nama, dun_kod))
    existing = cursor.fetchone()
    if existing:
        db.close()
        return {"success": True, "id": existing[0], "nama": nama, "existing": True}
    # Get dun_id
    cursor.execute("SELECT id FROM dun WHERE kod = ?", (dun_kod,))
    dun = cursor.fetchone()
    if not dun:
        db.close()
        raise HTTPException(status_code=400, detail=f"DUN {dun_kod} not found")
    dun_id = dun[0]
    cursor.execute("INSERT INTO pdm (dun_id, nama) VALUES (?, ?)", (dun_id, nama))
    db.commit()
    new_id = cursor.lastrowid
    db.close()
    return {"success": True, "id": new_id, "nama": nama}

@app.delete("/api/pdm/{pdm_nama}")
def delete_pdm(pdm_nama: str, data: PdmCreate = None, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    
    nama = pdm_nama.strip().upper()
    
    # Check if PDM exists
    cursor.execute("SELECT id FROM pdm WHERE nama = ?", (nama,))
    pdm = cursor.fetchone()
    if not pdm:
        db.close()
        raise HTTPException(status_code=404, detail=f"PDM {nama} tidak ditemui")
    
    pdm_id = pdm[0]
    
    # Check if any pengundi linked to this PDM (via dm field)
    cursor.execute("SELECT COUNT(*) FROM pengundi WHERE UPPER(dm) = ? AND status_fizikal = 'Hidup' AND status_rekod = 'Sah'", (nama,))
    pengundi_count = cursor.fetchone()[0]
    if pengundi_count > 0:
        db.close()
        raise HTTPException(status_code=400, detail=f"PDM {nama} masih mempunyai {pengundi_count} pengundi. Tidak boleh dipadam.")
    
    # Delete the PDM
    cursor.execute("DELETE FROM pdm WHERE id = ?", (pdm_id,))
    db.commit()
    db.close()
    
    return {"success": True, "message": f"PDM {nama} berjaya dipadamkan"}

# ===== ENDPOINT LOKALITI =====
@app.post("/api/lokaliti")
def create_lokaliti(request: Request, data: LokalitiCreate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin", "Petugas Padang"])
    db = get_db()
    cursor = db.cursor()
    nama = data.nama.strip().upper()
    dm = data.dm.strip().upper()
    # Check if Lokaliti already exists in same PDM (via kampung table)
    cursor.execute("""
        SELECT k.id FROM kampung k 
        JOIN pdm p ON p.id = k.pdm_id 
        WHERE k.nama = ? AND p.nama = ?
    """, (nama, dm))
    existing = cursor.fetchone()
    if existing:
        db.close()
        return {"success": True, "id": existing[0], "nama": nama, "existing": True}
    # Get pdm_id
    cursor.execute("SELECT id FROM pdm WHERE nama = ?", (dm,))
    pdm = cursor.fetchone()
    if not pdm:
        db.close()
        raise HTTPException(status_code=400, detail=f"PDM {dm} tidak ditemui")
    pdm_id = pdm[0]
    cursor.execute("INSERT INTO kampung (pdm_id, nama) VALUES (?, ?)", (pdm_id, nama))
    db.commit()
    new_id = cursor.lastrowid
    db.close()
    return {"success": True, "id": new_id, "nama": nama}

@app.delete("/api/lokaliti/{lokaliti_nama}")
def delete_lokaliti(lokaliti_nama: str, data: LokalitiCreate = None, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    
    nama = lokaliti_nama.strip().upper()
    
    # Check if Lokaliti exists in kampung table
    cursor.execute("SELECT id FROM kampung WHERE nama = ?", (nama,))
    kampung = cursor.fetchone()
    if not kampung:
        db.close()
        raise HTTPException(status_code=404, detail=f"Lokaliti {nama} tidak ditemui")
    
    kampung_id = kampung[0]
    
    # Check if any pengundi linked to this Lokaliti
    cursor.execute("SELECT COUNT(*) FROM pengundi WHERE UPPER(lokaliti) = ? AND status_fizikal = 'Hidup' AND status_rekod = 'Sah'", (nama,))
    pengundi_count = cursor.fetchone()[0]
    if pengundi_count > 0:
        db.close()
        raise HTTPException(status_code=400, detail=f"Lokaliti {nama} masih mempunyai {pengundi_count} pengundi. Tidak boleh dipadam.")
    
    # Delete the Lokaliti
    cursor.execute("DELETE FROM kampung WHERE id = ?", (kampung_id,))
    db.commit()
    db.close()
    
    return {"success": True, "message": f"Lokaliti {nama} berjaya dipadamkan"}

@app.get("/api/lokaliti")
def get_lokaliti_list(dun: Optional[str] = None, dm: Optional[str] = None, user=Depends(get_current_user)):
    """Pulangkan senarai lokaliti dari table pengundi secara langsung (single source of truth).
       Boleh filter oleh DUN (dun) dan/atau PDM (dm).
       POKA-YOKE: Jika tiada filter DUN/PDM, pulangkan SEMUA lokaliti tanpa status filter."""
    db = get_db()
    cursor = db.cursor()
    
    if dun and dun.strip() or dm and dm.strip():
        # Filtered mode: apply DUN/PDM + status constraints (Hidup & Sah)
        where_clauses = ["p.lokaliti IS NOT NULL AND p.lokaliti != ''", "p.status_fizikal = 'Hidup'", "p.status_rekod = 'Sah'"]
        params = []
        
        if dun and dun.strip():
            dun_val = dun.strip().upper()
            where_clauses.append("p.dun_id = (SELECT id FROM dun WHERE kod = ?)")
            params.append(dun_val)
        
        if dm and dm.strip():
            dm_val = dm.strip().upper()
            where_clauses.append("UPPER(p.dm) = ?")
            params.append(dm_val)
        
        where_sql = "WHERE " + " AND ".join(where_clauses)
        
        cursor.execute(f"""
            SELECT p.lokaliti AS nama, COUNT(p.id) AS jumlah_pengundi
            FROM pengundi p
            {where_sql}
            GROUP BY p.lokaliti
            ORDER BY p.lokaliti
        """, params)
    else:
        # Unfiltered mode (Poka-Yoke): ALL localities - no status filter
        cursor.execute("""
            SELECT p.lokaliti AS nama, COUNT(p.id) AS jumlah_pengundi
            FROM pengundi p
            WHERE p.lokaliti IS NOT NULL AND p.lokaliti != ''
            GROUP BY p.lokaliti
            ORDER BY p.lokaliti
        """)
    
    lokaliti = [{"nama": row[0], "jumlah_pengundi": row[1]} for row in cursor.fetchall()]
    db.close()
    return lokaliti

# ===== FUNGSI CHECK PERANAN =====
def check_peranan(user: dict, peranan_dibenarkan: list):
    if user["peranan"] not in peranan_dibenarkan:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Akses ditolak. Peranan '{user['peranan']}' tidak dibenarkan."
        )


# ===== ENDPOINT LOGIN (diimport dari secure_auth.py — JANGAN UBAH) =====
@app.post("/api/login")
def login(req: LoginRequest):
    return login_endpoint(req.username, req.kata_laluan)


# ===== ENDPOINT PENGUNDI =====

# Senarai PDM untuk dropdown — single source of truth dari master table pdm
@app.get("/api/pdm")
def get_pdm_list(user=Depends(get_current_user)):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT p.nama, p.dun_id, d.kod AS dun_kod,
               COUNT(pg.id) AS jumlah_pengundi
        FROM pdm p
        JOIN dun d ON d.id = p.dun_id
        LEFT JOIN pengundi pg ON UPPER(pg.dm) = UPPER(p.nama)
            AND pg.status_fizikal = 'Hidup'
            AND pg.status_rekod = 'Sah'
        GROUP BY p.nama, p.dun_id, d.kod
        ORDER BY p.nama
    """)
    pdms = [{"nama": row[0], "dun_id": row[1], "dun_kod": row[2], "jumlah_pengundi": row[3]} for row in cursor.fetchall()]
    db.close()
    return pdms

# Mapping PDM → DUN untuk fallback — pengundi lama belum ada dun_id
PDM_TO_DUN = {
    "BARU-BARU": "N12", "BATANGAN": "N12", "INDAI": "N12", "KINDU": "N12",
    "PENIMBAWAN": "N12", "SERUSUP": "N12", "TAMBALANG": "N12",
    "BERUNGIS": "N13", "GAYANG": "N13", "MARABAHAI": "N13", "MENGKABONG": "N13",
    "NONGKOULUD": "N13", "TELIPOK": "N13", "TUARAN BANDAR": "N13",
    "GAYARATAU": "N14", "KILANG BATA": "N14", "MENGKALADOI": "N14", "RANI": "N14",
    "RENGALIS": "N14", "RUNGUS": "N14", "SAWAH": "N14", "TAMPARULI": "N14",
    "TELIBONG": "N14", "TENGHILAN": "N14", "TOPOKON": "N14",
    "BONGOL": "N15", "KELAWAT": "N15", "KIULU": "N15", "MALANGANG": "N15",
    "MANTOB": "N15", "PAHU": "N15", "PORING": "N15", "PUKAK": "N15",
    "RANGALAU": "N15", "SIMPANGAN": "N15", "TAGINAMBUR": "N15",
    "TIONG SIMPODON": "N15", "TOGOP": "N15", "TOMIS": "N15", "TUDAN": "N15"
}

# Senarai DUN (sertakan jumlah pengundi) — single source of truth dari table pengundi
@app.get("/api/dun")
def get_dun_list(user=Depends(get_current_user)):
    db = get_db()
    cursor = db.cursor()
    
    # Poka-yoke: LEFT JOIN dengan table pengundi untuk count tepat (bukan guna PDM fallback)
    cursor.execute("""
        SELECT d.kod, d.nama, COUNT(p.id) AS jumlah_pengundi
        FROM dun d
        LEFT JOIN pengundi p ON p.dun_id = d.id
            AND p.status_fizikal = 'Hidup'
            AND p.status_rekod = 'Sah'
        GROUP BY d.kod, d.nama
        ORDER BY d.kod
    """)
    
    duns = [{"kod": row[0], "nama": row[0] + " " + row[1], "jumlah_pengundi": row[2]} for row in cursor.fetchall()]
    db.close()
    return duns

# Senarai PDM mengikut DUN — single source of truth dari master table pdm
@app.get("/api/pdm/dun/{dun_kod}")
def get_pdm_by_dun(dun_kod: str, user=Depends(get_current_user)):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT p.nama, COUNT(pg.id) AS jumlah_pengundi
        FROM pdm p
        JOIN dun d ON d.id = p.dun_id
        LEFT JOIN pengundi pg ON UPPER(pg.dm) = UPPER(p.nama)
            AND pg.status_fizikal = 'Hidup'
            AND pg.status_rekod = 'Sah'
        WHERE d.kod = ?
        GROUP BY p.nama
        ORDER BY p.nama
    """, (dun_kod,))
    pdms = [{"nama": row[0], "jumlah_pengundi": row[1]} for row in cursor.fetchall()]
    db.close()
    return pdms

# Dashboard stats per-PDM (untuk 4 PDM table dalam dashboard)
@app.get("/api/dashboard/pdm/{dun_kod}")
def get_dashboard_pdm(dun_kod: str, user=Depends(get_current_user)):
    """Pulangkan data per-PDM untuk sesuatu DUN — digunakan oleh 4 PDM table di frontend."""
    try:
        db = get_db()
        cursor = db.cursor()

        THN_SEMASA = 2026
        cursor.execute("""
            SELECT
                p.dm,
                COUNT(p.id) AS jumlah,
                SUM(CASE WHEN p.status_sokongan = 'Putih' THEN 1 ELSE 0 END) AS putih,
                SUM(CASE WHEN p.status_sokongan = 'Atas Pagar' THEN 1 ELSE 0 END) AS atas_pagar,
                SUM(CASE WHEN p.status_sokongan = 'Hitam' THEN 1 ELSE 0 END) AS hitam,
                SUM(CASE WHEN p.status_sokongan IS NULL OR p.status_sokongan NOT IN ('Putih', 'Atas Pagar', 'Hitam') THEN 1 ELSE 0 END) AS tidak_dikenali,
                SUM(CASE WHEN p.status_fizikal = 'Meninggal Dunia' THEN 1 ELSE 0 END) AS meninggal,
                SUM(CASE WHEN p.tahun_lahir IS NOT NULL AND (?) - p.tahun_lahir BETWEEN 18 AND 30 THEN 1 ELSE 0 END) AS usia_18_30,
                SUM(CASE WHEN p.tahun_lahir IS NOT NULL AND (?) - p.tahun_lahir BETWEEN 31 AND 59 THEN 1 ELSE 0 END) AS usia_31_59,
                SUM(CASE WHEN p.tahun_lahir IS NOT NULL AND (?) - p.tahun_lahir >= 60 THEN 1 ELSE 0 END) AS usia_60plus,
                COUNT(DISTINCT p.ketua_keluarga_id) AS jumlah_ketua_keluarga
            FROM pengundi p
            WHERE p.dun_id = (SELECT id FROM dun WHERE kod = ?)
              AND p.status_fizikal = 'Hidup'
              AND p.status_rekod = 'Sah'
              AND p.dm IS NOT NULL AND p.dm != ''
            GROUP BY p.dm
            ORDER BY p.dm
        """, (THN_SEMASA, THN_SEMASA, THN_SEMASA, dun_kod))

        data = []
        for row in cursor.fetchall():
            data.append({
                "dm": row["dm"],
                "jumlah": row["jumlah"],
                "putih": row["putih"],
                "atas_pagar": row["atas_pagar"],
                "hitam": row["hitam"],
                "tidak_dikenali": row["tidak_dikenali"],
                "meninggal": row["meninggal"],
                "usia_18_30": row["usia_18_30"],
                "usia_31_59": row["usia_31_59"],
                "usia_60plus": row["usia_60plus"],
                "jumlah_ketua_keluarga": row["jumlah_ketua_keluarga"]
            })

        db.close()
        return {"success": True, "dun_kod": dun_kod, "data": data}
    except Exception as e:
        import traceback
        print(f"❌ ERROR /api/dashboard/pdm/{dun_kod}: {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        # Fallback: pulangkan data kosong
        return {"success": False, "dun_kod": dun_kod, "data": [], "error": str(e)}


# Dashboard stats untuk DUN tertentu (filter ikut PDM dalam DUN)
@app.get("/api/dashboard/dun/{dun_kod}")
def get_dashboard_dun(request: Request, dun_kod: str, dm: Optional[str] = None, user=Depends(get_current_user)):
    try:
        db = get_db()
        cursor = db.cursor()

        where = "WHERE status_fizikal = 'Hidup' AND status_rekod = 'Sah' AND dun_id = (SELECT id FROM dun WHERE kod = ?)"
        params = [dun_kod]
        if dm:
            where += " AND dm = ?"
            params.append(dm)
        
        THN_SEMASA = 2026

        # Jumlah pengundi
        cursor.execute(f"SELECT COUNT(*) FROM pengundi p {where}", params)
        jumlah_pengundi = cursor.fetchone()[0]

        # Status sokongan
        cursor.execute(f"SELECT status_sokongan, COUNT(*) as jumlah FROM pengundi {where} GROUP BY status_sokongan ORDER BY jumlah DESC", params)
        sokongan = {}
        for row in cursor.fetchall():
            key = row["status_sokongan"] or "Tiada"
            sokongan[key] = row["jumlah"]

        db.close()

        return {
            "jumlah_pengundi": jumlah_pengundi,
            "sokongan": sokongan,
            "dun_kod": dun_kod
        }
    except Exception as e:
        import traceback
        print(f"⚠️ Dashboard DUN error (fallback to empty data): {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        # Fallback: pulangkan data kosong supaya frontend tidak crash
        return {
            "jumlah_pengundi": 0,
            "sokongan": {},
            "dun_kod": dun_kod,
            "error": str(e)
        }


# Dashboard stats - guna parameter dun (filter ikut DUN)
@app.get("/api/dashboard")
def get_dashboard(request: Request, dun: Optional[str] = None, user=Depends(get_current_user)):
    try:
        db = get_db()
        cursor = db.cursor()

        where = "WHERE status_fizikal = 'Hidup' AND status_rekod = 'Sah'"
        params = []
        if dun:
            where += " AND dun_id = (SELECT id FROM dun WHERE kod = ?)"
            params.append(dun)
        
        THN_SEMASA = 2026

        # Jumlah pengundi
        cursor.execute(f"SELECT COUNT(*) FROM pengundi {where}", params)
        jumlah_pengundi = cursor.fetchone()[0]

        # Status sokongan
        cursor.execute(f"SELECT status_sokongan, COUNT(*) as jumlah FROM pengundi {where} GROUP BY status_sokongan ORDER BY jumlah DESC", params)
        sokongan = {}
        for row in cursor.fetchall():
            key = row["status_sokongan"] or "Tiada"
            sokongan[key] = row["jumlah"]

        # Jantina
        cursor.execute(f"SELECT jantina, COUNT(*) as jumlah FROM pengundi {where} GROUP BY jantina", params)
        jantina = {}
        for row in cursor.fetchall():
            key = row["jantina"] or "Tidak Diketahui"
            jantina[key] = row["jumlah"]

        # Status fizikal
        cursor.execute(f"SELECT status_fizikal, COUNT(*) as jumlah FROM pengundi {where} GROUP BY status_fizikal", params)
        fizikal = {}
        for row in cursor.fetchall():
            key = row["status_fizikal"] or "Hidup"
            fizikal[key] = row["jumlah"]

        # Pengundi mengikut lokaliti
        cursor.execute(f"SELECT lokaliti, COUNT(*) as jumlah FROM pengundi {where} GROUP BY lokaliti ORDER BY jumlah DESC LIMIT 10", params)
        lokaliti_data = []
        for row in cursor.fetchall():
            lokaliti_data.append({"nama": row["lokaliti"] or "Tiada", "jumlah": row["jumlah"]})

        # Klasifikasi Umur
        cursor.execute(f"""
            SELECT 
                CASE 
                    WHEN (tahun_lahir IS NULL) THEN 'Tidak Diketahui'
                    WHEN ({THN_SEMASA} - tahun_lahir) BETWEEN 18 AND 30 THEN 'Belia'
                    WHEN ({THN_SEMASA} - tahun_lahir) BETWEEN 31 AND 59 THEN 'Dewasa'
                    WHEN ({THN_SEMASA} - tahun_lahir) >= 60 THEN 'Warga Emas'
                    ELSE 'Lain-lain'
                END as klasifikasi,
                COUNT(*) as jumlah,
                SUM(CASE WHEN jantina = 'L' THEN 1 ELSE 0 END) as lelaki,
                SUM(CASE WHEN jantina = 'P' THEN 1 ELSE 0 END) as perempuan
            FROM pengundi {where}
            GROUP BY klasifikasi
        """, params)
        klasifikasi_umur = {}
        for row in cursor.fetchall():
            klasifikasi_umur[row["klasifikasi"]] = {
                "jumlah": row["jumlah"],
                "lelaki": row["lelaki"],
                "perempuan": row["perempuan"]
            }

        # Umur purata
        cursor.execute(f"SELECT AVG({THN_SEMASA} - tahun_lahir) FROM pengundi {where} AND tahun_lahir IS NOT NULL", params)
        row = cursor.fetchone()
        purata_umur = round(row[0], 1) if row and row[0] is not None else 0

        # Sokongan mengikut klasifikasi umur
        cursor.execute(f"""
            SELECT 
                CASE 
                    WHEN (tahun_lahir IS NULL) THEN 'Tidak Diketahui'
                    WHEN ({THN_SEMASA} - tahun_lahir) BETWEEN 18 AND 30 THEN 'Belia'
                    WHEN ({THN_SEMASA} - tahun_lahir) BETWEEN 31 AND 59 THEN 'Dewasa'
                    WHEN ({THN_SEMASA} - tahun_lahir) >= 60 THEN 'Warga Emas'
                    ELSE 'Lain-lain'
                END as klasifikasi,
                COALESCE(status_sokongan, 'Tiada') as sokongan,
                COUNT(*) as jumlah
            FROM pengundi {where}
            GROUP BY klasifikasi, sokongan
            ORDER BY klasifikasi, sokongan
        """, params)
        sokongan_ikut_umur = {}
        for row in cursor.fetchall():
            k = row["klasifikasi"]
            s = row["sokongan"]
            j = row["jumlah"]
            if k not in sokongan_ikut_umur:
                sokongan_ikut_umur[k] = {}
            sokongan_ikut_umur[k][s] = j

        db.close()

        return {
            "jumlah_pengundi": jumlah_pengundi,
            "sokongan": sokongan,
            "jantina": jantina,
            "fizikal": fizikal,
            "lokaliti_teratas": lokaliti_data,
            "klasifikasi_umur": klasifikasi_umur,
            "purata_umur": purata_umur,
            "sokongan_ikut_umur": sokongan_ikut_umur
        }
    except Exception as e:
        import traceback
        error_detail = f"{type(e).__name__}: {str(e)}"
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail)


# Senarai pengundi dengan carian dan tapisan pelbagai (multi-select)
@app.get("/api/pengundi")
def get_pengundi(
    request: Request,
    page: int = 1,
    per_page: int = 50,
    search: Optional[str] = None,
    dm: Optional[str] = Query(None, alias="dm[]"),
    dun: Optional[str] = None,
    lokaliti: Optional[str] = Query(None, alias="lokaliti[]"),
    sokongan: Optional[str] = Query(None, alias="sokongan[]"),
    status_rekod: Optional[str] = None,
    ketua_keluarga: Optional[str] = Query(None, alias="ketua_keluarga[]"),
    pegawai_penyelaras: Optional[str] = Query(None, alias="pegawai_penyelaras[]"),
    user=Depends(get_current_user)
):
    db = get_db()
    cursor = db.cursor()

    where_parts = []
    params = []

    if search:
        where_parts.append("(UPPER(p.no_kp) LIKE UPPER(?) OR UPPER(p.nama_penuh) LIKE UPPER(?))")
        params.extend([f"%{search}%", f"%{search}%"])

    # Multi-select dm filter — tolak jika nilai kosong
    dm_list = []
    if dm and dm.strip():
        if isinstance(dm, str):
            dm_list = [d.strip() for d in dm.split(',') if d.strip()]
        elif isinstance(dm, list):
            dm_list = [d.strip() for d in dm if d.strip()]
    if dm_list:
        placeholders = ', '.join(['?'] * len(dm_list))
        where_parts.append(f"p.dm IN ({placeholders})")
        params.extend(dm_list)

    # DUN filter — filter by DUN code (p.dun_id -> dun.kod)
    if dun and dun.strip():
        where_parts.append("p.dun_id = (SELECT id FROM dun WHERE kod = ?)")
        params.append(dun.strip())

    # Multi-select lokaliti filter
    lokaliti_list = []
    if lokaliti:
        if isinstance(lokaliti, str):
            lokaliti_list = [l.strip() for l in lokaliti.split(',') if l.strip()]
        elif isinstance(lokaliti, list):
            lokaliti_list = [l.strip() for l in lokaliti if l.strip()]
    if lokaliti_list:
        placeholders = ', '.join(['?'] * len(lokaliti_list))
        where_parts.append(f"p.lokaliti IN ({placeholders})")
        params.extend(lokaliti_list)

    # Multi-select sokongan filter
    sokongan_list = []
    if sokongan:
        if isinstance(sokongan, str):
            sokongan_list = [s.strip() for s in sokongan.split(',') if s.strip()]
        elif isinstance(sokongan, list):
            sokongan_list = [s.strip() for s in sokongan if s.strip()]
    if sokongan_list:
        # Handle "Tiada" (NULL/empty) specially — SQL IN doesn't match NULL
        sokongan_vals = [s for s in sokongan_list if s != "Tiada"]
        ada_tiada = "Tiada" in sokongan_list
        sub_parts = []
        if sokongan_vals:
            placeholders = ', '.join(['?'] * len(sokongan_vals))
            sub_parts.append(f"p.status_sokongan IN ({placeholders})")
            params.extend(sokongan_vals)
        if ada_tiada:
            sub_parts.append("(p.status_sokongan IS NULL OR p.status_sokongan = '')")
        if sub_parts:
            where_parts.append("(" + " OR ".join(sub_parts) + ")")

    if status_rekod:
        where_parts.append("p.status_rekod = ?")
        params.append(status_rekod)

    # Multi-select ketua_keluarga filter (ID)
    ketua_list = []
    if ketua_keluarga:
        if isinstance(ketua_keluarga, str):
            ketua_list = [int(k.strip()) for k in ketua_keluarga.split(',') if k.strip().isdigit()]
        elif isinstance(ketua_keluarga, list):
            ketua_list = [int(k) for k in ketua_keluarga if str(k).strip().isdigit()]
    if ketua_list:
        placeholders = ', '.join(['?'] * len(ketua_list))
        where_parts.append(f"p.ketua_keluarga_id IN ({placeholders})")
        params.extend(ketua_list)

    # Multi-select pegawai_penyelaras filter (ID)
    pegawai_list = []
    if pegawai_penyelaras:
        if isinstance(pegawai_penyelaras, str):
            pegawai_list = [int(p.strip()) for p in pegawai_penyelaras.split(',') if p.strip().isdigit()]
        elif isinstance(pegawai_penyelaras, list):
            pegawai_list = [int(p) for p in pegawai_penyelaras if str(p).strip().isdigit()]
    if pegawai_list:
        placeholders = ', '.join(['?'] * len(pegawai_list))
        where_parts.append(f"p.pegawai_penyelaras_id IN ({placeholders})")
        params.extend(pegawai_list)

    where = ""
    if where_parts:
        where = "WHERE " + " AND ".join(where_parts)

    # Count total
    cursor.execute(f"SELECT COUNT(*) FROM pengundi p {where}", params)
    total = cursor.fetchone()[0]

    # Get page data
    offset = (page - 1) * per_page
    cursor.execute(f"""
        SELECT p.id, p.no_kp, p.nama_penuh, p.jantina, p.tahun_lahir, p.dm, p.lokaliti,
               p.no_telefon, p.status_sokongan, p.status_fizikal, p.adalah_pemilik_apps, p.status_rekod, p.sumber_pdm,
               p.ketua_keluarga_id, p.pegawai_penyelaras_id,
               kk.nama_penuh AS ketua_keluarga_nama,
               pp.nama_penuh AS pegawai_penyelaras_nama
        FROM pengundi p
        LEFT JOIN ketua_keluarga kk ON p.ketua_keluarga_id = kk.id
        LEFT JOIN pegawai_penyelaras pp ON p.pegawai_penyelaras_id = pp.id
        {where}
        ORDER BY p.id
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])

    pengundi = [dict(row) for row in cursor.fetchall()]
    
    # Count non-null values per column for the filtered dataset (poka-yoke: single source of truth)
    cursor.execute(f"""
        SELECT 
            COUNT(p.no_kp) AS cnt_no_kp,
            COUNT(p.nama_penuh) AS cnt_nama_penuh,
            COUNT(p.jantina) AS cnt_jantina,
            COUNT(p.tahun_lahir) AS cnt_tahun_lahir,
            COUNT(p.dm) AS cnt_dm,
            COUNT(p.lokaliti) AS cnt_lokaliti,
            COUNT(p.status_sokongan) AS cnt_sokongan,
            COUNT(kk.nama_penuh) AS cnt_ketua_keluarga,
            COUNT(pp.nama_penuh) AS cnt_pegawai_penyelaras
        FROM pengundi p
        LEFT JOIN ketua_keluarga kk ON p.ketua_keluarga_id = kk.id
        LEFT JOIN pegawai_penyelaras pp ON p.pegawai_penyelaras_id = pp.id
        {where}
    """, params)
    counts = cursor.fetchone()
    column_counts = {
        "no_kp": counts[0] if counts else 0,
        "nama_penuh": counts[1] if counts else 0,
        "jantina": counts[2] if counts else 0,
        "tahun_lahir": counts[3] if counts else 0,
        "dm": counts[4] if counts else 0,
        "lokaliti": counts[5] if counts else 0,
        "status_sokongan": counts[6] if counts else 0,
        "ketua_keluarga_nama": counts[7] if counts else 0,
        "pegawai_penyelaras_nama": counts[8] if counts else 0
    }
    db.close()

    # Log aktiviti
    dm_str = ','.join(dm_list) if dm_list else 'Semua'
    lokaliti_str = ','.join(lokaliti_list) if lokaliti_list else 'Semua'
    sokongan_str = ','.join(sokongan_list) if sokongan_list else 'Semua'
    if search:
        log_activity(request, user, "Carian Pengundi", f"Carian: '{search}' (P: {total} keputusan)", no_kp_terlibat=search)
    else:
        log_activity(request, user, "Lihat Senarai Pengundi", f"Filter: dm={dm_str}, lokaliti={lokaliti_str}, sokongan={sokongan_str}")

    return {
        "data": pengundi,
        "column_counts": column_counts,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page)
    }


# Endpoint untuk dapatkan senarai pilihan filter — boleh terima parameter untuk relevansi
@app.get("/api/pengundi/filter-options")
def get_filter_options(
    request: Request,
    dun: Optional[str] = None,
    dm: Optional[str] = Query(None, alias="dm[]"),
    lokaliti: Optional[str] = Query(None, alias="lokaliti[]"),
    sokongan: Optional[str] = Query(None, alias="sokongan[]"),
    user=Depends(get_current_user)
):
    db = get_db()
    cursor = db.cursor()

    # Build WHERE clause based on active filters
    where_clauses = ["1=1"]
    params = []

    if dun and dun.strip():
        where_clauses.append("p.dun_id = (SELECT id FROM dun WHERE kod = ?)")
        params.append(dun.strip())

    if dm and dm.strip():
        dm_list = [d.strip() for d in dm.split(',') if d.strip()]
        if dm_list:
            placeholders = ', '.join(['?'] * len(dm_list))
            where_clauses.append(f"p.dm IN ({placeholders})")
            params.extend(dm_list)

    if lokaliti and lokaliti.strip():
        lokaliti_list = [l.strip() for l in lokaliti.split(',') if l.strip()]
        if lokaliti_list:
            placeholders = ', '.join(['?'] * len(lokaliti_list))
            where_clauses.append(f"p.lokaliti IN ({placeholders})")
            params.extend(lokaliti_list)

    if sokongan and sokongan.strip():
        sokongan_list = [s.strip() for s in sokongan.split(',') if s.strip()]
        if sokongan_list:
            placeholders = ', '.join(['?'] * len(sokongan_list))
            where_clauses.append(f"p.status_sokongan IN ({placeholders})")
            params.extend(sokongan_list)

    where_str = " AND ".join(where_clauses)

    # Gunakan subquery untuk setiap filter supaya relevan dengan filter aktif
    cursor.execute(f"SELECT DISTINCT p.dm FROM pengundi p WHERE {where_str} AND p.dm IS NOT NULL AND p.dm != '' ORDER BY p.dm", params)
    pdm_list = [r[0] for r in cursor.fetchall()]

    cursor.execute(f"SELECT DISTINCT p.lokaliti FROM pengundi p WHERE {where_str} AND p.lokaliti IS NOT NULL AND p.lokaliti != '' ORDER BY p.lokaliti", params)
    lokaliti_list = [r[0] for r in cursor.fetchall()]

    # Sokongan filter options: include "Tiada" if there are records with NULL/empty status_sokongan
    cursor.execute(f"SELECT COUNT(*) FROM pengundi p WHERE {where_str} AND (p.status_sokongan IS NULL OR p.status_sokongan = '')", params)
    ada_tiada = cursor.fetchone()[0] > 0
    cursor.execute(f"SELECT DISTINCT p.status_sokongan FROM pengundi p WHERE {where_str} AND p.status_sokongan IS NOT NULL AND p.status_sokongan != '' ORDER BY p.status_sokongan", params)
    sokongan_list = [r[0] for r in cursor.fetchall()]
    if ada_tiada:
        sokongan_list.append("Tiada")

    # Ketua Keluarga & Pegawai Penyelaras — dari table rasmi (tidak perlu filter)
    cursor.execute("SELECT id, nama_penuh FROM ketua_keluarga ORDER BY nama_penuh")
    ketua_keluarga_list = [{"id": r[0], "nama": r[1]} for r in cursor.fetchall()]

    cursor.execute("SELECT id, nama_penuh FROM pegawai_penyelaras ORDER BY nama_penuh")
    pegawai_penyelaras_list = [{"id": r[0], "nama": r[1]} for r in cursor.fetchall()]

    db.close()
    return {
        "pdm": pdm_list,
        "lokaliti": lokaliti_list,
        "sokongan": sokongan_list,
        "ketua_keluarga": ketua_keluarga_list,
        "pegawai_penyelaras": pegawai_penyelaras_list
    }


# Endpoint carian pengundi untuk searchable dropdown (Ketua Keluarga/Pegawai Penyelaras)
@app.get("/api/pengundi/search")
def search_pengundi_dropdown(
    q: str = "",
    page: int = 1,
    per_page: int = 20,
    user=Depends(get_current_user)
):
    """Cari pengundi untuk dropdown pilihan — return id + nama_penuh sahaja."""
    db = get_db()
    cursor = db.cursor()
    
    params = [f"%{q}%", f"%{q}%"]
    offset = (page - 1) * per_page
    
    cursor.execute("""
        SELECT id, no_kp, nama_penuh, dm, lokaliti
        FROM pengundi
        WHERE (UPPER(nama_penuh) LIKE UPPER(?) OR UPPER(no_kp) LIKE UPPER(?))
          AND status_fizikal = 'Hidup' AND status_rekod = 'Sah'
        ORDER BY nama_penuh ASC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])
    
    results = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("""
        SELECT COUNT(*) FROM pengundi
        WHERE (UPPER(nama_penuh) LIKE UPPER(?) OR UPPER(no_kp) LIKE UPPER(?))
          AND status_fizikal = 'Hidup' AND status_rekod = 'Sah'
    """, params)
    total = cursor.fetchone()[0]
    
    db.close()
    return {"data": results, "total": total, "page": page, "per_page": per_page}


# Endpoint carian Ketua Keluarga untuk searchable dropdown — cari pengundi yang menjadi Ketua Keluarga
@app.get("/api/ketua-keluarga/search")
def search_ketua_keluarga_dropdown(
    q: str = "",
    page: int = 1,
    per_page: int = 200,
    user=Depends(get_current_user)
):
    """Cari Ketua Keluarga dari table rasmi untuk dropdown pilihan."""
    db = get_db()
    cursor = db.cursor()
    
    params = [f"%{q}%"]
    offset = (page - 1) * per_page
    
    cursor.execute("""
        SELECT id, id AS no_kp, nama_penuh, '' AS dm, '' AS lokaliti
        FROM ketua_keluarga
        WHERE UPPER(nama_penuh) LIKE UPPER(?)
        ORDER BY nama_penuh ASC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])
    
    results = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("""
        SELECT COUNT(*) FROM ketua_keluarga
        WHERE UPPER(nama_penuh) LIKE UPPER(?)
    """, params)
    total = cursor.fetchone()[0]
    
    db.close()
    return {"data": results, "total": total, "page": page, "per_page": per_page}


# Endpoint carian Pegawai Penyelaras untuk searchable dropdown
@app.get("/api/pegawai-penyelaras/search")
def search_pegawai_penyelaras_dropdown(
    q: str = "",
    page: int = 1,
    per_page: int = 200,
    user=Depends(get_current_user)
):
    """Cari Pegawai Penyelaras dari master table rasmi untuk dropdown pilihan."""
    db = get_db()
    cursor = db.cursor()
    
    params = [f"%{q}%"]
    offset = (page - 1) * per_page
    
    cursor.execute("""
        SELECT id, id AS no_kp, nama_penuh, '' AS dm, '' AS lokaliti
        FROM pegawai_penyelaras
        WHERE UPPER(nama_penuh) LIKE UPPER(?)
        ORDER BY nama_penuh ASC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])
    
    results = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("""
        SELECT COUNT(*) FROM pegawai_penyelaras
        WHERE UPPER(nama_penuh) LIKE UPPER(?)
    """, params)
    total = cursor.fetchone()[0]
    
    db.close()
    return {"data": results, "total": total, "page": page, "per_page": per_page}


# ===== ENDPOINT IMPORT EXCEL (Admin sahaja) =====

# Muat turun templat Excel kosong
@app.get("/api/pengundi/template")
def download_template(user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Pengundi"

    headers = ['NO KP', 'NAMA PENUH', 'JANTINA', 'TAHUN LAHIR', 'DM', 'LOKALITI',
               'NO TELEFON', 'PUTIH', 'ATAS PAGAR', 'HITAM', 'TAK KENAL', 'X DUNIA']
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # Set column widths
    widths = [15, 35, 10, 12, 20, 25, 15, 8, 10, 8, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=TEMPLAT_IMPORT_PENGUNDI.xlsx"}
    )


# Import data dari Excel
@app.post("/api/pengundi/import-excel")
def import_excel(request: Request, file: UploadFile = File(...), user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sila muat naik fail Excel (.xlsx atau .xls)")

    try:
        contents = file.file.read()
        df = pd.read_excel(BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca fail Excel: {str(e)}")

    # Normalise column names: uppercase, strip whitespace
    df.columns = [str(c).strip().upper() for c in df.columns]

    # Validate required columns
    required_cols = ['NO KP', 'NAMA PENUH', 'JANTINA', 'TAHUN LAHIR', 'DM', 'LOKALITI']
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Lajur wajib tidak dijumpai: {', '.join(missing)}. Lajur mesti termasuk: {', '.join(required_cols)}"
        )

    db = get_db()
    cursor = db.cursor()
    errors = []
    berjaya = 0

    insert_sql = """
        INSERT INTO pengundi
        (no_kp, nama_penuh, jantina, tahun_lahir, dm, lokaliti, no_telefon,
         status_sokongan, status_fizikal, adalah_pemilik_apps, status_rekod, sumber_pdm, dicipta_pada)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    now = datetime.now().isoformat()

    for idx, row in df.iterrows():
        try:
            # Process no_kp
            no_kp = row.get('NO KP')
            if pd.isna(no_kp) or not str(no_kp).strip():
                errors.append(f"Baris {idx+2}: No KP kosong")
                continue
            no_kp_str = str(no_kp).strip()
            # Remove dashes
            no_kp_str = no_kp_str.replace('-', '').replace(' ', '')
            # Take last 12 digits if longer
            digits_only = re.sub(r'\D', '', no_kp_str)
            if len(digits_only) >= 12:
                no_kp_str = digits_only[-12:]
            else:
                no_kp_str = digits_only

            # Process nama
            nama = row.get('NAMA PENUH')
            if pd.isna(nama) or not str(nama).strip():
                errors.append(f"Baris {idx+2}: Nama Penuh kosong")
                continue
            nama_str = str(nama).strip().upper()

            # Process jantina
            jantina_raw = row.get('JANTINA')
            jantina = None
            if pd.notna(jantina_raw):
                j = str(jantina_raw).strip().upper()
                if j in ['L', 'LELAKI', 'Lelaki']:
                    jantina = 'L'
                elif j in ['P', 'PEREMPUAN', 'Perempuan']:
                    jantina = 'P'

            # Process tahun_lahir
            tahun = row.get('TAHUN LAHIR')
            tahun_lahir = None
            if pd.notna(tahun):
                try:
                    tahun_lahir = int(float(str(tahun).strip()))
                except:
                    pass

            # Process DM
            dm = str(row.get('DM', '')).strip().upper() if pd.notna(row.get('DM')) else ''

            # Process lokaliti
            lokaliti = str(row.get('LOKALITI', '')).strip() if pd.notna(row.get('LOKALITI')) else None

            # Process no_telefon
            tel = row.get('NO TELEFON')
            no_telefon = str(tel).strip() if pd.notna(tel) else None

            # Process status sokongan
            status_sokongan = None
            for col_check, val in [('PUTIH', 'Putih'), ('ATAS PAGAR', 'Atas Pagar'),
                                    ('HITAM', 'Hitam'), ('TAK KENAL', 'Tidak Kenal'),
                                    ('P', 'Putih'), ('AP', 'Atas Pagar'), ('H', 'Hitam')]:
                if col_check in df.columns:
                    cell = row.get(col_check)
                    if pd.notna(cell):
                        cell_str = str(cell).strip()
                        if cell_str in ['1', '1.0', '1.00', 'TRUE', 'true', 'Y', 'y']:
                            status_sokongan = val
                            break

            # Process status fizikal
            status_fizikal = 'Hidup'
            for col_check in ['X DUNIA', 'MENINGGAL', 'MATI']:
                if col_check in df.columns:
                    cell = row.get(col_check)
                    if pd.notna(cell):
                        cell_str = str(cell).strip()
                        if cell_str in ['1', '1.0', '1.00', 'TRUE', 'true', 'Y', 'y']:
                            status_fizikal = 'Meninggal Dunia'
                            break

            values = (
                no_kp_str, nama_str, jantina, tahun_lahir,
                dm, lokaliti, no_telefon,
                status_sokongan, status_fizikal, 0,
                'Menunggu_Kelulusan', 'Import Excel', now
            )
            cursor.execute(insert_sql, values)
            berjaya += 1

        except Exception as e:
            errors.append(f"Baris {idx+2}: {str(e)}")

    db.commit()
    db.close()

    log_activity(request, user, "Import Excel",
                 f"Import Excel: {berjaya} berjaya, {len(errors)} gagal | Fail: {file.filename}")

    return {
        "berjaya": berjaya,
        "gagal": len(errors),
        "jumlah": berjaya + len(errors),
        "errors": errors[:20]  # Max 20 errors to avoid huge response
    }


# Dapatkan pengundi by ID (sertakan nama ketua_keluarga & pegawai_penyelaras)
@app.get("/api/pengundi/{pengundi_id}")
def get_pengundi_by_id(request: Request, pengundi_id: int, user=Depends(get_current_user)):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT p.*, 
               d.kod AS dun,
               kk.nama_penuh AS ketua_keluarga_nama, 
               pp.nama_penuh AS pegawai_penyelaras_nama
        FROM pengundi p
        LEFT JOIN dun d ON d.id = p.dun_id
        LEFT JOIN ketua_keluarga kk ON p.ketua_keluarga_id = kk.id
        LEFT JOIN pegawai_penyelaras pp ON p.pegawai_penyelaras_id = pp.id
        WHERE p.id = ?
    """, (pengundi_id,))
    p = cursor.fetchone()
    db.close()
    if not p:
        raise HTTPException(status_code=404, detail="Pengundi tidak ditemui")
    
    # Log aktiviti: lihat detail pengundi
    log_activity(request, user, "Lihat Detail Pengundi", f"Lihat detail ID {pengundi_id}: {p['nama_penuh']}", no_kp_terlibat=p['no_kp'])
    
    return dict(p)


# Tambah pengundi baru (Petugas Padang - status_rekod = Menunggu_Kelulusan)
@app.post("/api/pengundi")
def create_pengundi(request: Request, data: PengundiCreate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin", "Petugas Padang"])

    status_rekod = "Menunggu_Kelulusan" if user["peranan"] == "Petugas Padang" else "Sah"

    db = get_db()
    cursor = db.cursor()

    # If new pegawai_penyelaras name provided (not an existing ID), insert first
    pegawai_id = data.pegawai_penyelaras_id
    if data.pegawai_penyelaras_nama_baru and not pegawai_id:
        cursor.execute("INSERT INTO pegawai_penyelaras (nama_penuh) VALUES (?)", (data.pegawai_penyelaras_nama_baru.strip(),))
        db.commit()
        pegawai_id = cursor.lastrowid

    # If new ketua_keluarga name provided (not an existing ID), insert first
    ketua_id = data.ketua_keluarga_id
    if data.ketua_keluarga_nama_baru and not ketua_id:
        cursor.execute("INSERT INTO ketua_keluarga (nama_penuh) VALUES (?)", (data.ketua_keluarga_nama_baru.strip(),))
        db.commit()
        ketua_id = cursor.lastrowid

    # Resolve dun_kod -> dun_id (frontend sends DUN code like "N12")
    dun_id = None
    if data.dun:
        cursor.execute("SELECT id FROM dun WHERE kod = ?", (data.dun.strip().upper(),))
        dun_row = cursor.fetchone()
        if dun_row:
            dun_id = dun_row[0]

    cursor.execute("""
        INSERT INTO pengundi
        (no_kp, nama_penuh, jantina, tahun_lahir, dm, lokaliti, no_telefon,
         status_sokongan, status_fizikal, adalah_pemilik_apps, status_rekod, sumber_pdm, dicipta_pada,
         ketua_keluarga_id, pegawai_penyelaras_id, dun_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.no_kp, data.nama_penuh, data.jantina, data.tahun_lahir,
        data.dm, data.lokaliti, data.no_telefon,
        data.status_sokongan, data.status_fizikal, 0,
        status_rekod, f"Didaftar oleh {user['username']}", datetime.now().isoformat(),
        ketua_id, pegawai_id, dun_id
    ))
    db.commit()
    new_id = cursor.lastrowid
    db.close()

    # Log aktiviti
    log_activity(request, user, "Tambah Pengundi", 
                 f"Tambah pengundi baru: {data.nama_penuh} (KP: {data.no_kp}, PDM: {data.dm}, DUN: {data.dun}, status: {status_rekod})",
                 no_kp_terlibat=data.no_kp)

    return {"message": "Pengundi berjaya didaftarkan", "id": new_id, "status_rekod": status_rekod}


# Kemaskini pengundi
@app.put("/api/pengundi/{pengundi_id}")
def update_pengundi(request: Request, pengundi_id: int, data: PengundiUpdate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin", "Petugas Padang"])

    db = get_db()
    cursor = db.cursor()

    # Get existing
    cursor.execute("SELECT * FROM pengundi WHERE id = ?", (pengundi_id,))
    existing = cursor.fetchone()
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Pengundi tidak ditemui")

    # Build update — allow null for ketua_keluarga_id & pegawai_penyelaras_id
    update_fields = {}
    raw_data = data.model_dump(exclude_unset=True)
    for field, value in raw_data.items():
        if value is not None:
            update_fields[field] = value
        elif field in ('ketua_keluarga_id', 'pegawai_penyelaras_id'):
            # Explicit null: user wants to clear the field
            update_fields[field] = None

    if not update_fields and not any(f in raw_data for f in ('ketua_keluarga_id', 'pegawai_penyelaras_id')):
        db.close()
        return {"message": "Tiada perubahan dibuat"}

    # If Petugas Padang, set status to Menunggu_Kelulusan
    if user["peranan"] == "Petugas Padang":
        update_fields["status_rekod"] = "Menunggu_Kelulusan"

    set_clause = ", ".join([f"{k} = ?" for k in update_fields.keys()])
    values = list(update_fields.values()) + [pengundi_id]

    cursor.execute(f"UPDATE pengundi SET {set_clause} WHERE id = ?", values)
    db.commit()
    db.close()

    # Log aktiviti
    changed_fields = ', '.join(update_fields.keys())
    log_activity(request, user, "Edit Pengundi",
                 f"Edit pengundi ID {pengundi_id}: {existing['nama_penuh']} - field diubah: {changed_fields}",
                 no_kp_terlibat=existing['no_kp'])

    return {"message": "Pengundi berjaya dikemaskini", "fields_updated": list(update_fields.keys())}


# ===== DELETE PENGUNDI (Admin sahaja) - untuk padam terus rekod =====
@app.delete("/api/pengundi/{pengundi_id}")
def delete_pengundi(request: Request, pengundi_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    # Get data before delete
    cursor.execute("SELECT no_kp, nama_penuh, dm FROM pengundi WHERE id = ?", (pengundi_id,))
    p = cursor.fetchone()
    if not p:
        db.close()
        raise HTTPException(status_code=404, detail="Rekod tidak ditemui")

    no_kp = p["no_kp"]
    nama = p["nama_penuh"]
    dm = p["dm"]

    cursor.execute("DELETE FROM pengundi WHERE id = ?", (pengundi_id,))
    db.commit()
    db.close()

    log_activity(request, user, "Padam Pengundi",
                 f"Padam pengundi ID {pengundi_id}: {nama} (KP: {no_kp}, DM: {dm})",
                 no_kp_terlibat=no_kp)

    return {"message": f"Rekod {nama} berjaya dipadamkan"}


# ===== ENDPOINT APPROVAL QUEUE (Admin sahaja) =====
@app.get("/api/approval-queue")
def get_approval_queue(request: Request, page: int = 1, per_page: int = 50, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT COUNT(*) FROM pengundi WHERE status_rekod = 'Menunggu_Kelulusan'")
    total = cursor.fetchone()[0]

    offset = (page - 1) * per_page
    cursor.execute("""
        SELECT id, no_kp, nama_penuh, jantina, tahun_lahir, dm, lokaliti,
               no_telefon, status_sokongan, status_fizikal, status_rekod, sumber_pdm, dicipta_pada
        FROM pengundi
        WHERE status_rekod = 'Menunggu_Kelulusan'
        ORDER BY dicipta_pada DESC
        LIMIT ? OFFSET ?
    """, (per_page, offset))

    queue = [dict(row) for row in cursor.fetchall()]
    db.close()

    log_activity(request, user, "Lihat Approval Queue", f"Lihat queue kelulusan ({total} menunggu)")

    return {
        "data": queue,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page)
    }


# Luluskan rekod (Admin sahaja)
@app.post("/api/approval-queue/{pengundi_id}/lulus")
def approve_record(request: Request, pengundi_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()
    
    # Get data before approve
    cursor.execute("SELECT no_kp, nama_penuh FROM pengundi WHERE id = ?", (pengundi_id,))
    p = cursor.fetchone()
    if not p:
        db.close()
        raise HTTPException(status_code=404, detail="Rekod tidak ditemui")
    
    cursor.execute("UPDATE pengundi SET status_rekod = 'Sah' WHERE id = ? AND status_rekod = 'Menunggu_Kelulusan'", (pengundi_id,))
    if cursor.rowcount == 0:
        db.close()
        raise HTTPException(status_code=404, detail="Rekod sudah diluluskan")
    db.commit()
    db.close()

    log_activity(request, user, "Lulus Rekod",
                 f"Luluskan rekod ID {pengundi_id}: {p['nama_penuh']}",
                 no_kp_terlibat=p['no_kp'])

    return {"message": "Rekod berjaya diluluskan"}


# Tolak rekod (Admin sahaja)
@app.delete("/api/approval-queue/{pengundi_id}/tolak")
def reject_record(request: Request, pengundi_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()
    
    # Get data before reject
    cursor.execute("SELECT no_kp, nama_penuh FROM pengundi WHERE id = ? AND status_rekod = 'Menunggu_Kelulusan'", (pengundi_id,))
    p = cursor.fetchone()
    if not p:
        db.close()
        raise HTTPException(status_code=404, detail="Rekod tidak ditemui atau sudah diproses")
    
    no_kp = p['no_kp']
    nama = p['nama_penuh']
    
    cursor.execute("DELETE FROM pengundi WHERE id = ? AND status_rekod = 'Menunggu_Kelulusan'", (pengundi_id,))
    db.commit()
    db.close()

    log_activity(request, user, "Tolak Rekod",
                 f"Tolak rekod ID {pengundi_id}: {nama}",
                 no_kp_terlibat=no_kp)

    return {"message": "Rekod berjaya ditolak dan dipadamkan"}


# ===== APPROVE ALL PENDING (Admin sahaja) =====
@app.post("/api/pengundi/approve-all")
def approve_all_pending(request: Request, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    # Count pending records
    cursor.execute("SELECT COUNT(*) FROM pengundi WHERE status_rekod = 'Menunggu_Kelulusan'")
    count = cursor.fetchone()[0]

    if count == 0:
        db.close()
        return {"message": "Tiada rekod yang menunggu kelulusan", "jumlah": 0}

    # Bulk approve all pending
    cursor.execute("UPDATE pengundi SET status_rekod = 'Sah' WHERE status_rekod = 'Menunggu_Kelulusan'")
    db.commit()
    db.close()

    log_activity(request, user, "Luluskan Semua",
                 f"{count} rekod diluluskan secara pukal")

    return {"message": f"{count} rekod berjaya diluluskan", "jumlah": count}


# ===== ENDPOINT AUDIT LOGS (Admin sahaja) =====
@app.get("/api/audit-logs")
def get_audit_logs(
    request: Request,
    page: int = 1,
    per_page: int = 50,
    username: Optional[str] = None,
    tindakan: Optional[str] = None,
    search: Optional[str] = None,
    user=Depends(get_current_user)
):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    where_parts = []
    params = []

    if username:
        where_parts.append("username = ?")
        params.append(username)
    if tindakan:
        where_parts.append("tindakan = ?")
        params.append(tindakan)
    if search:
        where_parts.append("(penerangan LIKE ? OR no_kp_terlibat LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])

    where = ""
    if where_parts:
        where = "WHERE " + " AND ".join(where_parts)

    cursor.execute(f"SELECT COUNT(*) FROM audit_logs {where}", params)
    total = cursor.fetchone()[0]

    offset = (page - 1) * per_page
    cursor.execute(f"""
        SELECT id, user_id, username, peranan, tindakan, penerangan, no_kp_terlibat, endpoint, ip_address, user_agent, dicipta_pada
        FROM audit_logs {where}
        ORDER BY dicipta_pada DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])

    logs = [dict(row) for row in cursor.fetchall()]
    db.close()

    return {
        "data": logs,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page)
    }


# ===== ENDPOINT PENGURUSAN PENGGUNA (Admin sahaja) =====
@app.get("/api/users")
def get_users(user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, username, nama_penuh, peranan, dm, aktif, dicipta_pada FROM users ORDER BY id")
    users = [dict(row) for row in cursor.fetchall()]
    db.close()
    return users


@app.post("/api/users")
def create_user(request: Request, data: PenggunaCreate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    # Check if username exists
    cursor.execute("SELECT id FROM users WHERE username = ?", (data.username,))
    if cursor.fetchone():
        db.close()
        raise HTTPException(status_code=400, detail="Nama pengguna sudah wujud")

    cursor.execute("""
        INSERT INTO users (username, nama_penuh, kata_laluan, peranan, dm, aktif, dicipta_pada)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        data.username, data.nama_penuh, hash_kata_laluan(data.kata_laluan),
        data.peranan, data.dm, 1, datetime.now().isoformat()
    ))
    db.commit()
    db.close()

    log_activity(request, user, "Tambah Pengguna",
                 f"Tambah pengguna baru: {data.username} ({data.peranan})")

    return {"message": f"Pengguna '{data.username}' berjaya dicipta"}


@app.patch("/api/users/{user_id}")
def patch_user(request: Request, user_id: int, data: dict = Body(...), user=Depends(get_current_user)):
    """Kemaskini status pengguna (aktif/nyahaktif)."""
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, username, nama_penuh FROM users WHERE id = ?", (user_id,))
    existing = cursor.fetchone()
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemui")
    # Only allow updating 'aktif' field
    if "aktif" in data:
        cursor.execute("UPDATE users SET aktif = ? WHERE id = ?", (1 if data["aktif"] else 0, user_id))
        db.commit()
        status_str = "diaktifkan" if data["aktif"] else "dinyahaktifkan"
        log_activity(request, user, f"Pengguna {status_str}",
                     f"User ID {user_id} ({existing['username']}) {status_str}")
    db.close()
    return {"message": "Pengguna berjaya dikemaskini"}


@app.delete("/api/users/{user_id}")
def delete_user(request: Request, user_id: int, user=Depends(get_current_user)):
    """Padam pengguna (Admin sahaja). Tidak boleh padam diri sendiri."""
    check_peranan(user, ["Admin"])
    if user.get("user_id") == user_id:
        raise HTTPException(status_code=400, detail="Tidak boleh memadamkan akaun sendiri")
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, username, nama_penuh FROM users WHERE id = ?", (user_id,))
    existing = cursor.fetchone()
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemui")
    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    db.close()
    log_activity(request, user, "Padam Pengguna",
                 f"Padam pengguna ID {user_id}: {existing['username']} ({existing['nama_penuh']})")
    return {"message": f"Pengguna '{existing['username']}' berjaya dipadamkan"}


# ===== SURVEY MODELS =====
class SurveyCreate(BaseModel):
    title: str
    description: str = ""
    questions: str  # JSON string of questions array

class SurveySubmit(BaseModel):
    survey_id: int
    answers: str  # JSON string
    respondent_info: str = ""


# ===== SURVEY ENDPOINTS =====

# Mock AI: generate questions based on topic
TEMPLATE_QUESTIONS = {
    "air": [
        {"id": "q1", "type": "short_text", "question": "Adakah anda menghadapi masalah bekalan air bersih?", "required": True},
        {"id": "q2", "type": "multiple_choice", "question": "Berapa kerap masalah air berlaku?", "options": ["Setiap hari", "Beberapa kali seminggu", "Sekali seminggu", "Jarang"], "required": True},
        {"id": "q3", "type": "checkboxes", "question": "Apakah jenis masalah air yang anda hadapi?", "options": ["Air keruh", "Air berbau", "Tekanan rendah", "Bekalan terputus", "Lain-lain"], "required": False},
        {"id": "q4", "type": "short_text", "question": "Apakah cadangan anda untuk menambah baik bekalan air?", "required": False},
    ],
    "jalan": [
        {"id": "q1", "type": "short_text", "question": "Adakah jalan di kawasan anda dalam keadaan baik?", "required": True},
        {"id": "q2", "type": "multiple_choice", "question": "Apakah masalah jalan utama?", "options": ["Jalan berlubang", "Jalan sempit", "Tiada lampu jalan", "Longkang tersumbat"], "required": True},
        {"id": "q3", "type": "short_text", "question": "Lokasi mana yang paling memerlukan perhatian?", "required": True},
    ],
    "sampah": [
        {"id": "q1", "type": "multiple_choice", "question": "Berapa kerap kutipan sampah di kawasan anda?", "options": ["Setiap hari", "2-3 kali seminggu", "Seminggu sekali", "Jarang/Tidak pernah"], "required": True},
        {"id": "q2", "type": "checkboxes", "question": "Apakah isu berkaitan sampah?", "options": ["Kutipan tidak tetap", "Tiada tong sampah", "Pembakaran terbuka", "Longkang tersumbat"], "required": False},
    ],
    "kesihatan": [
        {"id": "q1", "type": "multiple_choice", "question": "Adakah anda berpuas hati dengan perkhidmatan kesihatan?", "options": ["Sangat puas", "Puas", "Kurang puas", "Tidak puas"], "required": True},
        {"id": "q2", "type": "checkboxes", "question": "Apakah kemudahan kesihatan yang diperlukan?", "options": ["Klinik bergerak", "Farmasi", "Ambulans", "Program kesihatan"], "required": False},
    ],
}

DEFAULT_QUESTIONS = [
    {"id": "q1", "type": "short_text", "question": "Apakah pendapat anda tentang isu ini?", "required": True},
    {"id": "q2", "type": "multiple_choice", "question": "Sejauh manakah isu ini memberi kesan kepada anda?", "options": ["Sangat terkesan", "Terkesan", "Kurang terkesan", "Tidak terkesan"], "required": True},
    {"id": "q3", "type": "short_text", "question": "Ada sebarang cadangan?", "required": False},
]


@app.post("/api/surveys/generate")
def generate_survey(prompt_data: dict = Body(...), user=Depends(get_current_user)):
    """Mock AI: generate questions based on topic text"""
    prompt = prompt_data.get("prompt", "").lower()

    # Find matching template or use default
    questions = DEFAULT_QUESTIONS
    for keyword, qs in TEMPLATE_QUESTIONS.items():
        if keyword in prompt:
            questions = qs
            break

    return {
        "title": f"Kajian: {prompt.capitalize() if prompt else 'Tinjauan'}",
        "description": f"Soal selidik berkaitan {prompt if prompt else 'topik'}",
        "questions": questions
    }


@app.get("/api/surveys")
def get_surveys(user=Depends(get_current_user)):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT s.*, COUNT(sr.id) as response_count
        FROM Survey s
        LEFT JOIN SurveyResponse sr ON sr.survey_id = s.id
        GROUP BY s.id
        ORDER BY s.created_at DESC
    """)
    surveys = [dict(row) for row in cursor.fetchall()]
    db.close()
    return surveys


@app.post("/api/surveys")
def create_survey(request: Request, data: SurveyCreate, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    now = datetime.now().isoformat()
    cursor.execute(
        "INSERT INTO Survey (title, description, questions, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
        (data.title, data.description, data.questions, user.get("user_id"), now)
    )
    db.commit()
    survey_id = cursor.lastrowid
    db.close()
    log_activity(request, user, "Buat Survey", f"Survey '{data.title}' dicipta")
    return {"id": survey_id, "message": "Survey berjaya dicipta"}


@app.get("/api/surveys/{survey_id}")
def get_survey(survey_id: int, user=Depends(get_current_user)):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM Survey WHERE id = ?", (survey_id,))
    survey = cursor.fetchone()
    if not survey:
        db.close()
        raise HTTPException(status_code=404, detail="Survey tidak ditemui")
    survey = dict(survey)
    # Parse questions JSON
    try:
        survey["questions"] = json.loads(survey["questions"])
    except:
        pass
    db.close()
    return survey


@app.post("/api/surveys/submit")
def submit_survey_response(request: Request, data: SurveySubmit):
    """Public endpoint - no auth required for respondents"""
    db = get_db()
    cursor = db.cursor()
    # Verify survey exists
    cursor.execute("SELECT id FROM Survey WHERE id = ?", (data.survey_id,))
    if not cursor.fetchone():
        db.close()
        raise HTTPException(status_code=404, detail="Survey tidak ditemui")
    now = datetime.now().isoformat()
    cursor.execute(
        "INSERT INTO SurveyResponse (survey_id, answers, respondent_info, submitted_at) VALUES (?, ?, ?, ?)",
        (data.survey_id, data.answers, data.respondent_info, now)
    )
    db.commit()
    db.close()
    return {"message": "Respons berjaya dihantar"}


@app.get("/api/surveys/{survey_id}/responses")
def get_survey_responses(survey_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT id, answers, respondent_info, submitted_at FROM SurveyResponse WHERE survey_id = ? ORDER BY submitted_at DESC", (survey_id,))
    responses = [dict(row) for row in cursor.fetchall()]
    # Parse answers JSON
    for r in responses:
        try:
            r["answers"] = json.loads(r["answers"])
        except:
            pass
    db.close()
    return responses


@app.delete("/api/surveys/{survey_id}")
def delete_survey(request: Request, survey_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    cursor.execute("DELETE FROM SurveyResponse WHERE survey_id = ?", (survey_id,))
    cursor.execute("DELETE FROM Survey WHERE id = ?", (survey_id,))
    db.commit()
    db.close()
    log_activity(request, user, "Padam Survey", f"Survey ID {survey_id} dipadamkan")
    return {"message": "Survey dipadamkan"}



# Dashboard Ringkasan Parlimen — data agregat untuk 4 DUN
@app.get("/api/dashboard/ringkasan")
def get_dashboard_ringkasan(user=Depends(get_current_user)):
    """Pulangkan data ringkasan untuk setiap 4 DUN."""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute("""
            SELECT 
                d.kod, d.nama,
                COUNT(p.id) AS jumlah_berdaftar,
                SUM(CASE WHEN p.status_sokongan = 'Putih' THEN 1 ELSE 0 END) AS putih,
                SUM(CASE WHEN p.status_sokongan = 'Atas Pagar' THEN 1 ELSE 0 END) AS atas_pagar,
                SUM(CASE WHEN p.status_sokongan = 'Hitam' THEN 1 ELSE 0 END) AS hitam,
                SUM(CASE WHEN p.status_sokongan IS NULL OR p.status_sokongan NOT IN ('Putih', 'Atas Pagar', 'Hitam') THEN 1 ELSE 0 END) AS tidak_dikenali,
                SUM(CASE WHEN p.status_fizikal = 'Meninggal Dunia' THEN 1 ELSE 0 END) AS meninggal,
                COUNT(DISTINCT p.ketua_keluarga_id) AS jumlah_ketua_keluarga
            FROM pengundi p
            JOIN dun d ON p.dun_id = d.id
            WHERE d.parlimen_id = (SELECT id FROM parlimen WHERE kod = 'P170')
            GROUP BY d.kod, d.nama
            ORDER BY d.kod
        """)
        
        dun_list = []
        for row in cursor.fetchall():
            dun_list.append({
                "dun_kod": row["kod"],
                "dun_nama": f"DUN {row['kod']} {row['nama']}",
                "jumlah_berdaftar": row["jumlah_berdaftar"],
                "putih": row["putih"],
                "atas_pagar": row["atas_pagar"],
                "hitam": row["hitam"],
                "tidak_dikenali": row["tidak_dikenali"],
                "meninggal": row["meninggal"],
                "jumlah_ketua_keluarga": row["jumlah_ketua_keluarga"]
            })
        
        cursor.execute("""
            SELECT 
                COUNT(*) AS jumlah_berdaftar,
                SUM(CASE WHEN status_sokongan = 'Putih' THEN 1 ELSE 0 END) AS putih,
                SUM(CASE WHEN status_sokongan = 'Atas Pagar' THEN 1 ELSE 0 END) AS atas_pagar,
                SUM(CASE WHEN status_sokongan = 'Hitam' THEN 1 ELSE 0 END) AS hitam,
                SUM(CASE WHEN status_sokongan IS NULL OR status_sokongan NOT IN ('Putih', 'Atas Pagar', 'Hitam') THEN 1 ELSE 0 END) AS tidak_dikenali,
                COUNT(DISTINCT ketua_keluarga_id) AS jumlah_ketua_keluarga
            FROM pengundi
        """)
        total = cursor.fetchone()
        
        db.close()

        return {
            "success": True,
            "parlimen": "P170 Tuaran",
            "dun": dun_list,
            "jumlah_keseluruhan": {
                "jumlah_berdaftar": total["jumlah_berdaftar"],
                "putih": total["putih"],
                "atas_pagar": total["atas_pagar"],
                "hitam": total["hitam"],
                "tidak_dikenali": total["tidak_dikenali"],
                "jumlah_ketua_keluarga": total["jumlah_ketua_keluarga"]
            }
        }
    except Exception as e:
        import traceback
        print(f"❌ ERROR /api/dashboard/ringkasan: {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)}")


# ===== DIAGNOSTIC DASHBOARD TEST =====
@app.get("/api/dashboard-test")
def dashboard_test(user=Depends(get_current_user)):
    """Test endpoint untuk diagnosis dashboard."""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM pengundi WHERE status_fizikal = 'Hidup' AND status_rekod = 'Sah'")
        count = cursor.fetchone()[0]
        cursor.execute("SELECT jantina, COUNT(*) FROM pengundi WHERE status_fizikal = 'Hidup' AND status_rekod = 'Sah' GROUP BY jantina")
        jantina = {r[0] or 'Unknown': r[1] for r in cursor.fetchall()}
        cursor.execute("SELECT status_sokongan, COUNT(*) FROM pengundi WHERE status_fizikal = 'Hidup' AND status_rekod = 'Sah' GROUP BY status_sokongan")
        sokongan = {r[0] or 'Tiada': r[1] for r in cursor.fetchall()}
        db.close()
        return {"success": True, "count": count, "jantina": jantina, "sokongan": sokongan, "msg": "Dashboard test OK"}
    except Exception as e:
        return {"success": False, "error": str(e), "error_type": type(e).__name__}


# ============================================================
# PENGURUSAN PEGAWAI PENYELARAS (Admin only)
# ============================================================

class PegawaiPenyelarasCreate(BaseModel):
    nama_penuh: str
    no_kp: str
    no_telefon: str
    dm: Optional[str] = None  # Optional PDM/DUN reference tag

class PegawaiPenyelarasUpdate(BaseModel):
    nama_penuh: Optional[str] = None
    no_kp: Optional[str] = None
    no_telefon: Optional[str] = None
    dm: Optional[str] = None

# Senarai Pegawai Penyelaras (Admin only)
@app.get("/api/pegawai-penyelaras/list")
def get_pegawai_penyelaras_list(
    request: Request,
    search: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
    user=Depends(get_current_user)
):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()

    where_parts = ["pp.aktif = 1"]
    params = []

    if search:
        where_parts.append("(UPPER(pp.nama_penuh) LIKE UPPER(?) OR UPPER(pp.no_kp) LIKE UPPER(?))")
        params.extend([f"%{search}%", f"%{search}%"])

    where = "WHERE " + " AND ".join(where_parts)

    # Count total
    cursor.execute(f"SELECT COUNT(*) FROM pegawai_penyelaras pp {where}", params)
    total = cursor.fetchone()[0]

    offset = (page - 1) * per_page
    cursor.execute(f"""
        SELECT pp.id, pp.nama_penuh, pp.no_kp, pp.no_telefon, pp.dm, pp.dicipta_pada,
               COALESCE((SELECT COUNT(*) FROM pengundi p 
                         WHERE p.pegawai_penyelaras_id = pp.id 
                         AND p.status_fizikal = 'Hidup' 
                         AND p.status_rekod = 'Sah'), 0) AS jumlah_pengundi,
               d.kod AS dun_kod,
               d.nama AS dun_nama
        FROM pegawai_penyelaras pp
        LEFT JOIN dun d ON d.id = pp.dun_id
        {where}
        ORDER BY pp.nama_penuh
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])

    pegawai = [dict(row) for row in cursor.fetchall()]
    db.close()

    log_activity(request, user, "Lihat Senarai Pegawai Penyelaras",
                 f"Senarai Pegawai Penyelaras ({total} rekod)")

    return {
        "data": pegawai,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page)
    }


# KPI Stats untuk Pengurusan Pegawai Penyelaras
@app.get("/api/pegawai-penyelaras/stats")
def get_pegawai_penyelaras_stats(user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()

    # Total Pegawai Penyelaras aktif
    cursor.execute("SELECT COUNT(*) FROM pegawai_penyelaras WHERE aktif = 1")
    total_pegawai = cursor.fetchone()[0]

    # DUN coverage - jumlah DUN yang mempunyai sekurang-kurangnya 1 pegawai
    cursor.execute("""
        SELECT COUNT(DISTINCT d.kod) 
        FROM pegawai_penyelaras pp
        JOIN dun d ON d.id = pp.dun_id
        WHERE pp.aktif = 1 AND pp.dun_id IS NOT NULL
    """)
    dun_covered = cursor.fetchone()[0]

    # Total DUN
    cursor.execute("SELECT COUNT(*) FROM dun")
    total_dun = cursor.fetchone()[0]

    # PDM coverage - jumlah PDM dengan pegawai
    cursor.execute("""
        SELECT COUNT(DISTINCT pp.dm) 
        FROM pegawai_penyelaras pp
        WHERE pp.aktif = 1 AND pp.dm IS NOT NULL AND pp.dm != ''
    """)
    pdm_covered = cursor.fetchone()[0]

    # Total PDM
    cursor.execute("SELECT COUNT(*) FROM pdm")
    total_pdm = cursor.fetchone()[0]

    # Total pengundi yang terikat dengan mana-mana pegawai
    cursor.execute("""
        SELECT COUNT(DISTINCT p.id) 
        FROM pengundi p
        JOIN pegawai_penyelaras pp ON pp.id = p.pegawai_penyelaras_id
        WHERE pp.aktif = 1 AND p.status_fizikal = 'Hidup' AND p.status_rekod = 'Sah'
    """)
    pengundi_terikat = cursor.fetchone()[0]

    db.close()

    return {
        "total_pegawai": total_pegawai,
        "dun_coverage": f"{dun_covered}/{total_dun}",
        "dun_covered": dun_covered,
        "total_dun": total_dun,
        "pdm_coverage": f"{pdm_covered}/{total_pdm}",
        "pdm_covered": pdm_covered,
        "total_pdm": total_pdm,
        "pengundi_terikat": pengundi_terikat
    }


# Dapatkan Pegawai Penyelaras by ID
@app.get("/api/pegawai-penyelaras/{pegawai_id}")
def get_pegawai_penyelaras_by_id(pegawai_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT pp.*, d.kod AS dun_kod, d.nama AS dun_nama,
               COALESCE((SELECT COUNT(*) FROM pengundi p 
                         WHERE p.pegawai_penyelaras_id = pp.id 
                         AND p.status_fizikal = 'Hidup' 
                         AND p.status_rekod = 'Sah'), 0) AS jumlah_pengundi
        FROM pegawai_penyelaras pp
        LEFT JOIN dun d ON d.id = pp.dun_id
        WHERE pp.id = ? AND pp.aktif = 1
    """, (pegawai_id,))
    pegawai = cursor.fetchone()
    db.close()
    if not pegawai:
        raise HTTPException(status_code=404, detail="Pegawai Penyelaras tidak ditemui")
    return dict(pegawai)


# Daftar Pegawai Penyelaras baru
@app.post("/api/pegawai-penyelaras")
def create_pegawai_penyelaras(
    request: Request, 
    data: PegawaiPenyelarasCreate, 
    user=Depends(get_current_user)
):
    check_peranan(user, ["Admin"])
    
    # Validate required fields
    if not data.nama_penuh or not data.nama_penuh.strip():
        raise HTTPException(status_code=400, detail="Nama Penuh wajib diisi")
    if not data.no_kp or not data.no_kp.strip():
        raise HTTPException(status_code=400, detail="No Kad Pengenalan wajib diisi")
    if not data.no_telefon or not data.no_telefon.strip():
        raise HTTPException(status_code=400, detail="No Telefon wajib diisi")

    db = get_db()
    cursor = db.cursor()

    # Check if no_kp already exists
    cursor.execute("SELECT id FROM pegawai_penyelaras WHERE no_kp = ? AND aktif = 1", (data.no_kp.strip(),))
    if cursor.fetchone():
        db.close()
        raise HTTPException(status_code=400, detail=f"No KP {data.no_kp} sudah didaftarkan untuk pegawai lain")

    # Resolve DUN from dm if provided
    dun_id = None
    dm_value = data.dm.strip().upper() if data.dm and data.dm.strip() else None
    if dm_value:
        # Try to match DUN via PDM name first
        cursor.execute("""
            SELECT d.id FROM dun d
            JOIN pdm p ON p.dun_id = d.id
            WHERE UPPER(p.nama) = ?
            LIMIT 1
        """, (dm_value,))
        dun_row = cursor.fetchone()
        if dun_row:
            dun_id = dun_row[0]
        else:
            # Try to match as DUN code directly
            cursor.execute("SELECT id FROM dun WHERE kod = ?", (dm_value,))
            dun_row = cursor.fetchone()
            if dun_row:
                dun_id = dun_row[0]

    now = datetime.now().isoformat()
    cursor.execute("""
        INSERT INTO pegawai_penyelaras (nama_penuh, no_kp, no_telefon, dm, dun_id, aktif, dicipta_pada, dikemaskini_pada)
        VALUES (?, ?, ?, ?, ?, 1, ?, ?)
    """, (data.nama_penuh.strip().upper(), data.no_kp.strip(), data.no_telefon.strip(),
          dm_value, dun_id, now, now))
    db.commit()
    new_id = cursor.lastrowid
    db.close()

    log_activity(request, user, "Tambah Pegawai Penyelaras",
                 f"Tambah pegawai baru: {data.nama_penuh} (KP: {data.no_kp}, Telefon: {data.no_telefon})")

    return {"message": "Pegawai Penyelaras berjaya didaftarkan", "id": new_id}


# Kemaskini Pegawai Penyelaras
@app.put("/api/pegawai-penyelaras/{pegawai_id}")
def update_pegawai_penyelaras(
    request: Request, 
    pegawai_id: int, 
    data: PegawaiPenyelarasUpdate, 
    user=Depends(get_current_user)
):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    # Get existing
    cursor.execute("SELECT * FROM pegawai_penyelaras WHERE id = ?", (pegawai_id,))
    existing = cursor.fetchone()
    if not existing:
        db.close()
        raise HTTPException(status_code=404, detail="Pegawai Penyelaras tidak ditemui")

    # Build update fields
    update_fields = {}
    raw_data = data.model_dump(exclude_unset=True)
    for field, value in raw_data.items():
        if value is not None:
            if field == "no_kp":
                # Check uniqueness if changed
                cursor.execute("SELECT id FROM pegawai_penyelaras WHERE no_kp = ? AND id != ? AND aktif = 1",
                              (value.strip(), pegawai_id))
                if cursor.fetchone():
                    db.close()
                    raise HTTPException(status_code=400, detail=f"No KP {value} sudah digunakan oleh pegawai lain")
                update_fields[field] = value.strip()
            elif field == "nama_penuh":
                update_fields[field] = value.strip().upper()
            elif field == "no_telefon":
                update_fields[field] = value.strip()
            elif field == "dm":
                dm_val = value.strip().upper()
                update_fields[field] = dm_val
                # Try to resolve DUN
                cursor.execute("""
                    SELECT d.id FROM dun d
                    JOIN pdm p ON p.dun_id = d.id
                    WHERE UPPER(p.nama) = ?
                    LIMIT 1
                """, (dm_val,))
                dun_row = cursor.fetchone()
                if dun_row:
                    update_fields["dun_id"] = dun_row[0]
                else:
                    cursor.execute("SELECT id FROM dun WHERE kod = ?", (dm_val,))
                    dun_row = cursor.fetchone()
                    if dun_row:
                        update_fields["dun_id"] = dun_row[0]

    if not update_fields:
        db.close()
        return {"message": "Tiada perubahan dibuat"}

    update_fields["dikemaskini_pada"] = datetime.now().isoformat()

    set_clause = ", ".join([f"{k} = ?" for k in update_fields.keys()])
    values = list(update_fields.values()) + [pegawai_id]

    cursor.execute(f"UPDATE pegawai_penyelaras SET {set_clause} WHERE id = ?", values)
    db.commit()
    db.close()

    log_activity(request, user, "Edit Pegawai Penyelaras",
                 f"Edit pegawai ID {pegawai_id}: {existing['nama_penuh']} - field diubah: {', '.join(update_fields.keys())}")

    return {"message": "Pegawai Penyelaras berjaya dikemaskini", "fields_updated": list(update_fields.keys())}


# Padam Pegawai Penyelaras (soft delete - set aktif=0)
@app.delete("/api/pegawai-penyelaras/{pegawai_id}")
def delete_pegawai_penyelaras(request: Request, pegawai_id: int, user=Depends(get_current_user)):
    check_peranan(user, ["Admin"])

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT id, nama_penuh, no_kp FROM pegawai_penyelaras WHERE id = ? AND aktif = 1", (pegawai_id,))
    pegawai = cursor.fetchone()
    if not pegawai:
        db.close()
        raise HTTPException(status_code=404, detail="Pegawai Penyelaras tidak ditemui atau sudah tidak aktif")

    # Soft delete: set aktif = 0
    now = datetime.now().isoformat()
    cursor.execute("UPDATE pegawai_penyelaras SET aktif = 0, dikemaskini_pada = ? WHERE id = ?", (now, pegawai_id))
    
    # Unlink all pengundi that reference this pegawai (set to NULL)
    cursor.execute("UPDATE pengundi SET pegawai_penyelaras_id = NULL WHERE pegawai_penyelaras_id = ?", (pegawai_id,))
    
    db.commit()
    db.close()

    log_activity(request, user, "Padam Pegawai Penyelaras",
                 f"Padam pegawai ID {pegawai_id}: {pegawai['nama_penuh']} (KP: {pegawai['no_kp']}) - {cursor.rowcount} pengundi dinyahpaut")

    return {"message": f"Pegawai '{pegawai['nama_penuh']}' berjaya dipadamkan"}


# ===== PPU (PETUNJUK PRESTASI UTAMA) ENDPOINT =====
@app.get("/api/p_pegawai-penyelaras")
@app.get("/api/ppu/pegawai-penyelaras")
def get_ppu_pegawai_penyelaras(user=Depends(get_current_user)):
    """Pulangkan data agregat PPU — prestasi Pegawai Penyelaras berdasarkan data sebenar dari database."""
    try:
        db = get_db()
        cursor = db.cursor()

        # ============================================================
        # FIX: DUN MENGUNDI (2026-07-17)
        # 
        # Punca bug: Kod lama guna dm (PDM) utk cari DUN melalui
        # subquery pengundi — rantaian tidak stabil & lambat.
        #
        # SOLUSI MUTLAK (Database First):
        # JOIN penuh pegawai_penyelaras → pengundi → dun → parlimen
        # menggunakan pengundi_id sebagai kunci utama.
        # 
        # Table pegawai_penyelaras ada pengundi_id (FK ke pengundi)
        # Table pengundi ada dun_id (FK ke dun) & parlimen_id (FK ke parlimen)
        # ============================================================

        cursor.execute("""
            SELECT
                pp.id,
                pp.nama_penuh,
                pp.dm AS kawasan,
                -- DUN info: cuba JOIN via pengundi_id dulu; jika NULL, guna dm (PDM) untuk cari DUN
                COALESCE(
                    (SELECT d_out.kod || ' ' || d_out.nama FROM pengundi p_dm
                     LEFT JOIN dun d_out ON d_out.id = p_dm.dun_id
                     WHERE p_dm.id = pp.pengundi_id),
                    (SELECT d_out.kod || ' ' || d_out.nama FROM pengundi p_dm2
                     LEFT JOIN dun d_out ON d_out.id = p_dm2.dun_id
                     WHERE p_dm2.dm = pp.dm AND p_dm2.dun_id IS NOT NULL LIMIT 1),
                    '-'
                ) AS dun_mengundi,
                -- Parlimen: P170 Tuaran (semua DUN di bawah Parlimen yang sama)
                'P170 Tuaran' AS parlimen_mengundi,
                pp.dm AS pdm_nama,
                -- Statistik Rekrut K.K
                COALESCE((SELECT COUNT(*) FROM pengundi p2
                           WHERE p2.dm = pp.dm
                             AND p2.status_fizikal = 'Hidup'
                             AND p2.status_rekod = 'Sah'
                             AND p2.ketua_keluarga_id IS NOT NULL), 0) AS rekrut_kk,
                -- Statistik Rekrut Putih
                COALESCE((SELECT COUNT(*) FROM pengundi p2
                           WHERE p2.dm = pp.dm
                             AND p2.status_fizikal = 'Hidup'
                             AND p2.status_rekod = 'Sah'
                             AND p2.status_sokongan = 'Putih'), 0) AS rekrut_putih,
                -- Jumlah Pengundi dalam PDM
                COALESCE((SELECT COUNT(*) FROM pengundi p2
                           WHERE p2.dm = pp.dm
                             AND p2.status_fizikal = 'Hidup'
                             AND p2.status_rekod = 'Sah'), 0) AS jumlah_pengundi
            FROM pegawai_penyelaras pp
            ORDER BY pp.nama_penuh
        """)

        results = []
        for row in cursor.fetchall():
            results.append({
                "id": row["id"],
                "nama": row["nama_penuh"],
                "kawasan": row["kawasan"],
                "parlimen": row["parlimen_mengundi"],
                "dun": row["dun_mengundi"],
                "pdm": row["pdm_nama"] or '-',
                "rekrut_kk": row["rekrut_kk"] or 0,
                "rekrut_putih": row["rekrut_putih"] or 0,
                "jumlah_pengundi": row["jumlah_pengundi"] or 0
            })

        db.close()
        return {"data": results}
    except Exception as e:
        import traceback
        print(f"❌ ERROR /api/ppu/pegawai-penyelaras: {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)}")


# ===== CATCH-ALL: Static Files & SPA Fallback =====
# Diletakkan SELEPAS semua API routes supaya tidak mencuri request API
from fastapi.responses import FileResponse, HTMLResponse
import mimetypes

@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    """
    Serve static files & SPA fallback.
    - Paths bermula /api/* sudah ditangkap oleh route di atas
    - Static files (JS, CSS, PNG, etc.) — serve terus jika wujud
    - SPA routes — fallback ke index.html
    """
    file_path = os.path.join(STATIC_ROOT, full_path) if full_path else os.path.join(STATIC_ROOT, "index.html")
    
    # If it's a file that exists, serve it
    if os.path.isfile(file_path):
        content_type, _ = mimetypes.guess_type(file_path)
        return FileResponse(file_path, media_type=content_type or "application/octet-stream")
    
    # SPA fallback: serve index.html for any non-API route
    index_path = os.path.join(STATIC_ROOT, "index.html")
    if os.path.isfile(index_path):
        return FileResponse(index_path, media_type="text/html")
    
    # Ultimate fallback
    return HTMLResponse("<h1>404 Not Found</h1>", status_code=404)


# ===== ROOT =====
@app.get("/api/status")
def root():
    return {
        "app": "Sistem Pengurusan Pengundi Parlimen P170 Tuaran",
        "versi": "1.1.0",
        "status": "beroperasi",
        "fitur_baru": "Audit Trail (PDPA), Edit Pengundi"
    }


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
